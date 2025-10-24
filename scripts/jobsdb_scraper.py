import sys
import json
import time
import pickle
import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 配置常量
COOKIES_FILE = "jobsdb_cookies.pkl"
EXCEL_OUTPUT_DIR = "scraped_data"
JOBSDB_URL = "https://hk.jobsdb.com"
# 正确的登录 URL
LOGIN_URL = "https://hk.jobsdb.com/login"

def setup_driver():
    """初始化 Chrome WebDriver（使用自动驱动管理）"""
    options = webdriver.ChromeOptions()
    
    # 开发环境：保持浏览器可见以便调试
    # 生产环境：取消注释下面两行启用无头模式
    # options.add_argument('--headless=new')
    # options.add_argument('--window-size=1920,1080')
    
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    # 设置用户代理
    options.add_argument('--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.set_page_load_timeout(30)
        return driver
    except Exception as e:
        print(json.dumps({
            "status": "error",
            "message": f"浏览器初始化失败: {str(e)}"
        }), file=sys.stderr)
        raise

def save_cookies(driver, filepath=COOKIES_FILE):
    """保存 cookies 到文件"""
    try:
        with open(filepath, 'wb') as f:
            pickle.dump(driver.get_cookies(), f)
        print(json.dumps({
            "status": "info",
            "message": "Cookies 已保存"
        }))
        sys.stdout.flush()
    except Exception as e:
        print(json.dumps({
            "status": "warning",
            "message": f"保存 Cookies 失败: {str(e)}"
        }))
        sys.stdout.flush()

def load_cookies(driver, filepath=COOKIES_FILE):
    """从文件加载 cookies"""
    if not os.path.exists(filepath):
        return False
    try:
        driver.get(JOBSDB_URL)
        time.sleep(2)
        
        with open(filepath, 'rb') as f:
            cookies = pickle.load(f)
            for cookie in cookies:
                try:
                    driver.add_cookie(cookie)
                except:
                    continue
        return True
    except Exception as e:
        print(json.dumps({
            "status": "warning",
            "message": f"加载 Cookies 失败: {str(e)}"
        }))
        sys.stdout.flush()
        return False

def wait_for_login_completion(driver, timeout=300):
    """
    等待用户完成邮箱验证并登录
    检测多个可能的登录成功标志
    """
    print(json.dumps({
        "status": "waiting_verification",
        "message": "请检查邮箱并点击验证链接完成登录"
    }))
    sys.stdout.flush()
    
    start_time = time.time()
    check_interval = 2  # 每2秒检查一次
    
    while time.time() - start_time < timeout:
        try:
            current_url = driver.current_url
            
            # 检查 URL 是否已经离开登录页面
            if "login" not in current_url.lower() and "oauth" not in current_url.lower():
                # 尝试多种方式确认登录成功
                login_indicators = [
                    # 用户菜单/头像
                    "[data-cy='user-menu']",
                    "[data-automation='user-menu']",
                    "button[aria-label*='Account']",
                    "button[aria-label*='Profile']",
                    # 常见的已登录元素
                    ".user-avatar",
                    ".profile-menu",
                    "[class*='UserMenu']",
                    "[class*='userMenu']",
                    # 通过检查是否有"登出"按钮
                    "a[href*='logout']",
                    "button[data-cy*='logout']"
                ]
                
                for selector in login_indicators:
                    try:
                        driver.find_element(By.CSS_SELECTOR, selector)
                        print(json.dumps({
                            "status": "login_success",
                            "message": "登录成功！"
                        }))
                        sys.stdout.flush()
                        return True
                    except NoSuchElementException:
                        continue
                
                # 如果URL变了但找不到登录标志，等待一下再检查
                time.sleep(3)
                
            time.sleep(check_interval)
            
            # 每30秒提醒一次
            elapsed = time.time() - start_time
            if int(elapsed) % 30 == 0 and elapsed > 0:
                remaining = int(timeout - elapsed)
                print(json.dumps({
                    "status": "waiting",
                    "message": f"仍在等待登录... (剩余 {remaining} 秒)"
                }))
                sys.stdout.flush()
                
        except Exception as e:
            time.sleep(check_interval)
            continue
    
    return False

def login_to_jobsdb(driver, email):
    """
    处理 JobsDB 登录流程
    """
    try:
        # 1. 先尝试加载已保存的 cookies
        if load_cookies(driver):
            driver.get(JOBSDB_URL)
            time.sleep(3)
            
            # 检查是否已登录（通过查找登录按钮的缺失来判断）
            try:
                # 如果找到登录按钮，说明未登录
                driver.find_element(By.LINK_TEXT, "Sign in")
                print(json.dumps({
                    "status": "info",
                    "message": "已保存的 session 已过期，需要重新登录"
                }))
                sys.stdout.flush()
            except NoSuchElementException:
                # 找不到登录按钮，说明已经登录
                print(json.dumps({
                    "status": "login_success",
                    "message": "使用保存的 session 登录成功"
                }))
                sys.stdout.flush()
                return True
        
        # 2. 需要重新登录
        print(json.dumps({
            "status": "login_required",
            "message": "正在打开登录页面..."
        }))
        sys.stdout.flush()
        
        driver.get(LOGIN_URL)
        time.sleep(3)
        
        # 3. 查找并填写邮箱输入框
        # JobsDB 可能使用不同的输入框标识
        email_selectors = [
            "input[type='email']",
            "input[name='email']",
            "input[id*='email']",
            "input[placeholder*='email']",
            "input[placeholder*='Email']",
            "input[autocomplete='email']"
        ]
        
        email_input = None
        for selector in email_selectors:
            try:
                email_input = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                if email_input:
                    break
            except:
                continue
        
        if not email_input:
            raise Exception("找不到邮箱输入框，页面结构可能已更改")
        
        print(json.dumps({
            "status": "info",
            "message": "找到邮箱输入框，正在填写..."
        }))
        sys.stdout.flush()
        
        email_input.clear()
        time.sleep(1)
        email_input.send_keys(email)
        time.sleep(2)
        
        # 4. 查找并点击提交按钮
        button_selectors = [
            "button[type='submit']",
            "button[data-cy='submit']",
            "button[data-automation='submitButton']",
            "input[type='submit']",
            "button:contains('Continue')",
            "button:contains('Log in')"
        ]
        
        submit_button = None
        for selector in button_selectors:
            try:
                submit_button = driver.find_element(By.CSS_SELECTOR, selector)
                if submit_button and submit_button.is_displayed():
                    break
            except:
                continue
        
        if not submit_button:
            raise Exception("找不到提交按钮，页面结构可能已更改")
        
        print(json.dumps({
            "status": "info",
            "message": "正在提交邮箱..."
        }))
        sys.stdout.flush()
        
        submit_button.click()
        time.sleep(3)
        
        # 5. 等待邮箱验证
        if wait_for_login_completion(driver):
            save_cookies(driver)
            return True
        else:
            raise Exception("登录超时，请确认已点击邮箱验证链接")
            
    except TimeoutException:
        raise Exception("页面加载超时，请检查网络连接")
    except Exception as e:
        raise Exception(f"登录失败: {str(e)}")

def scrape_application_history(driver):
    """
    爬取投递记录
    """
    applications = []
    
    try:
        print(json.dumps({
            "status": "scraping",
            "message": "正在导航到申请记录页面..."
        }))
        sys.stdout.flush()
        
        # 方法1: 尝试直接访问申请历史页面的常见URL
        possible_urls = [
            f"{JOBSDB_URL}/profile/applications",
            f"{JOBSDB_URL}/my-activity/applied-jobs",
            f"{JOBSDB_URL}/applications",
            f"{JOBSDB_URL}/profile/job-applications",
        ]
        
        for url in possible_urls:
            try:
                driver.get(url)
                time.sleep(3)
                
                # 检查页面是否有效（不是404或重定向到登录）
                if "login" not in driver.current_url.lower() and "404" not in driver.page_source:
                    print(json.dumps({
                        "status": "info",
                        "message": f"成功访问: {url}"
                    }))
                    sys.stdout.flush()
                    break
            except:
                continue
        
        # 方法2: 如果直接访问失败，尝试通过导航菜单找到申请记录
        try:
            # 点击用户菜单
            menu_selectors = [
                "[data-cy='user-menu']",
                "[data-automation='user-menu']",
                "button[aria-label*='Account']",
                ".user-avatar"
            ]
            
            for selector in menu_selectors:
                try:
                    menu = driver.find_element(By.CSS_SELECTOR, selector)
                    menu.click()
                    time.sleep(2)
                    
                    # 查找"我的申请"或类似链接
                    applications_links = driver.find_elements(By.XPATH, 
                        "//a[contains(text(), 'Applications') or contains(text(), 'My applications') or contains(text(), '我的申请')]")
                    
                    if applications_links:
                        applications_links[0].click()
                        time.sleep(3)
                        break
                except:
                    continue
        except:
            pass
        
        print(json.dumps({
            "status": "scraping",
            "message": "正在分析页面结构..."
        }))
        sys.stdout.flush()
        
        # 等待页面加载 - 尝试多个可能的选择器
        page_loaded = False
        possible_container_selectors = [
            "[data-automation='job-list']",
            "[data-cy='job-list']",
            ".job-card",
            ".application-item",
            "[class*='ApplicationCard']",
            "[class*='JobCard']",
            "[class*='application-card']",
            "article",
            "[role='article']"
        ]
        
        for selector in possible_container_selectors:
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                page_loaded = True
                print(json.dumps({
                    "status": "info",
                    "message": f"找到内容容器: {selector}"
                }))
                sys.stdout.flush()
                break
            except:
                continue
        
        if not page_loaded:
            # 保存页面HTML用于调试
            with open("jobsdb_page_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            
            raise Exception("无法找到申请记录容器，页面HTML已保存到 jobsdb_page_debug.html 供调试")
        
        # 滚动加载所有内容
        print(json.dumps({
            "status": "scraping",
            "message": "正在加载所有申请记录..."
        }))
        sys.stdout.flush()
        
        last_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scrolls = 10
        
        while scroll_attempts < max_scrolls:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            new_height = driver.execute_script("return document.body.scrollHeight")
            
            if new_height == last_height:
                break
            
            last_height = new_height
            scroll_attempts += 1
        
        # 查找所有职位卡片
        job_card_selectors = [
            "[data-automation='job-card']",
            "[data-cy='job-card']",
            ".job-card",
            ".application-item",
            "[class*='ApplicationCard']",
            "[class*='JobCard']",
            "article[class*='application']",
            "div[class*='application-card']"
        ]
        
        job_cards = []
        for selector in job_card_selectors:
            job_cards = driver.find_elements(By.CSS_SELECTOR, selector)
            if job_cards:
                print(json.dumps({
                    "status": "info",
                    "message": f"找到 {len(job_cards)} 个职位卡片（使用选择器: {selector}）"
                }))
                sys.stdout.flush()
                break
        
        if not job_cards:
            raise Exception("找不到任何职位卡片，可能没有申请记录或页面结构已变更")
        
        # 提取每个职位的信息
        for idx, card in enumerate(job_cards):
            try:
                # 职位名称 - 尝试多个选择器
                title = None
                title_selectors = [
                    "h3", "h2", "h4",
                    ".job-title",
                    "[data-automation='job-title']",
                    "[class*='JobTitle']",
                    "[class*='title']",
                    "a[class*='title']"
                ]
                
                for selector in title_selectors:
                    try:
                        title_elem = card.find_element(By.CSS_SELECTOR, selector)
                        title = title_elem.text.strip()
                        if title:
                            break
                    except:
                        continue
                
                if not title:
                    continue  # 跳过无法获取标题的卡片
                
                # 公司名称
                company = "未知公司"
                company_selectors = [
                    ".company-name",
                    "[data-automation='company-name']",
                    "[class*='CompanyName']",
                    "[class*='company']",
                    "span[class*='company']"
                ]
                
                for selector in company_selectors:
                    try:
                        company_elem = card.find_element(By.CSS_SELECTOR, selector)
                        company = company_elem.text.strip()
                        if company:
                            break
                    except:
                        continue
                
                # 投递日期
                date = "未知"
                date_selectors = [
                    "time",
                    ".date",
                    ".apply-date",
                    "[class*='date']",
                    "span[class*='date']"
                ]
                
                for selector in date_selectors:
                    try:
                        date_elem = card.find_element(By.CSS_SELECTOR, selector)
                        date = date_elem.text.strip()
                        if date:
                            break
                    except:
                        continue
                
                # 状态
                status = "已投递"
                status_selectors = [
                    ".status",
                    "[data-automation='status']",
                    "[class*='Status']",
                    "span[class*='status']",
                    ".badge"
                ]
                
                for selector in status_selectors:
                    try:
                        status_elem = card.find_element(By.CSS_SELECTOR, selector)
                        status = status_elem.text.strip()
                        if status:
                            break
                    except:
                        continue
                
                # 职位链接
                link = ""
                try:
                    link_elem = card.find_element(By.CSS_SELECTOR, "a")
                    link = link_elem.get_attribute("href")
                except:
                    pass
                
                applications.append({
                    "title": title,
                    "company": company,
                    "date": date,
                    "status": status,
                    "link": link
                })
                
                if (idx + 1) % 10 == 0:
                    print(json.dumps({
                        "status": "info",
                        "message": f"已解析 {idx + 1}/{len(job_cards)} 条记录..."
                    }))
                    sys.stdout.flush()
                
            except Exception as e:
                # 跳过无法解析的卡片
                continue
        
        if not applications:
            # 保存页面用于调试
            with open("jobsdb_page_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            
            raise Exception("未能解析任何投递记录，页面HTML已保存到 jobsdb_page_debug.html")
        
        print(json.dumps({
            "status": "success",
            "message": f"成功爬取 {len(applications)} 条申请记录"
        }))
        sys.stdout.flush()
        
        return applications
        
    except TimeoutException:
        raise Exception("页面加载超时，请检查网络连接")
    except Exception as e:
        raise Exception(f"爬取失败: {str(e)}")

def save_to_excel(data, filename=None):
    """将数据保存为 Excel 文件"""
    if not os.path.exists(EXCEL_OUTPUT_DIR):
        os.makedirs(EXCEL_OUTPUT_DIR)
    
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"jobsdb_applications_{timestamp}.xlsx"
    
    filepath = os.path.join(EXCEL_OUTPUT_DIR, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "投递记录"
    
    # 设置表头
    headers = ["职位名称", "公司名称", "投递日期", "状态", "职位链接"]
    ws.append(headers)
    
    # 美化表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    for item in data:
        ws.append([
            item.get("title", ""),
            item.get("company", ""),
            item.get("date", ""),
            item.get("status", ""),
            item.get("link", "")
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    
    wb.save(filepath)
    
    # 返回绝对路径
    return os.path.abspath(filepath)

def main_scrape_logic():
    """主爬虫逻辑"""
    driver = None
    
    try:
        # 从命令行参数获取邮箱
        email = sys.argv[1] if len(sys.argv) > 1 else None
        
        if not email:
            raise Exception("未提供邮箱地址")
        
        print(json.dumps({
            "status": "init",
            "message": f"开始爬取流程，使用邮箱: {email}"
        }))
        sys.stdout.flush()
        
        # 初始化浏览器
        driver = setup_driver()
        
        # 登录
        login_to_jobsdb(driver, email)
        
        # 爬取数据
        applications = scrape_application_history(driver)
        
        # 保存到 Excel
        excel_path = save_to_excel(applications)
        
        return {
            "success": True,
            "count": len(applications),
            "data": applications,
            "excel_path": excel_path,
            "message": f"成功爬取 {len(applications)} 条记录"
        }
        
    except Exception as e:
        raise Exception(f"爬取过程出错: {str(e)}")
    
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass

if __name__ == "__main__":
    try:
        result = main_scrape_logic()
        print(json.dumps(result, ensure_ascii=False))
        sys.stdout.flush()
        
    except Exception as e:
        error_msg = str(e)
        print(json.dumps({
            "error": error_msg,
            "status": "failed"
        }, ensure_ascii=False), file=sys.stderr)
        sys.stderr.flush()
        sys.exit(1)